from models.pdf_manager import PDF
from models.dxf_manager import DXF
from models.dwg_manager import DWG
import openpyxl
import rarfile
import zipfile
import shutil
import os

os.environ["PATH"] += os.pathsep + "/usr/bin"

class ArquivoGrupo:
    """Representa um grupo de arquivos de uma mesma pasta."""
    def __init__(self, pasta):
        self.pasta = pasta
        self.pdfs = []
        self.dxfs = []
        self.dwgs = []
        self.rows_to_add_first = []
        self.rows_to_add = []

    def adicionar_arquivo(self, arquivo_obj):
        """Adiciona objetos aos grupos relevantes."""
        if isinstance(arquivo_obj, PDF):
            self.pdfs.append(arquivo_obj)
        elif isinstance(arquivo_obj, DXF):
            self.dxfs.append(arquivo_obj)
        elif isinstance(arquivo_obj, DWG):
            self.dwgs.append(arquivo_obj)

    def __repr__(self):
        return (f"Grupo: {self.pasta}\n"
                f"PDFs: {len(self.pdfs)} objetos\n"
                f"DXFs: {len(self.dxfs)} objetos\n"
                f"DWGs: {len(self.dwgs)} objetos\n")


# Processa o arquivo .zip de entrada e os 'transforma' em um arquivo.zip processado e uma planilha
class ZipFolderManager:
    
    def __init__(self, zip_folder):
        
        self.pdfs = []
        self.dxfs = []
        self.dwgs = []
        self.grupos = []
        
        self.folder = zip_folder
        
        self.extraction_folder = "DESENHOS PDFs"
        self.organization_folder = 'OrganizedFiles'
        self.organization_folder_zip = 'OrganizedFilesZip'
        self.all_pdfs = 'OrganizedFiles/PDFs'
        self.all_dwgs = 'OrganizedFiles/DWGs'
        self.all_dxfs = 'OrganizedFiles/DXFs'
        self.descompact_folder = ZipFolderManager.extract_folder(self)
        self.agroup_designs()
        self.sheets = ZipFolderManager.create_sheet(self)
        self.organize_folders()
        self.processed_zip = ZipFolderManager.zip_file_process(self)
            
# Recebe a pasta .zip e realiza a primeira descompactação              
    def extract_folder(self): 
                
                # Remove a pasta de extração se ela já existir
                if os.path.exists(self.extraction_folder):
                    shutil.rmtree(self.extraction_folder)
                
                # Cria a pasta de extração
                os.makedirs(self.extraction_folder)

                # Verifica o tipo de arquivo e realiza a extração
                if self.folder.lower().endswith('.zip'):
                    with zipfile.ZipFile(self.folder, 'r') as archive:
                        archive.extractall(self.extraction_folder)
                        
                elif self.folder.lower().endswith('.rar'):
                    with rarfile.RarFile(self.folder, 'r') as archive:
                        archive.extractall(self.extraction_folder)
                        
                # Chama a função para extrair subarquivos
                sub = self.extraction_folder
                extrair = extract_subfiles(sub)
                return extrair
 
    def agroup_designs(self):

        for root, dirs, files in os.walk(self.descompact_folder):
            # Ignora pastas que contêm outras subpastas
            if dirs:
                continue

            grupo = ArquivoGrupo(root)
            arquivos_encontrados = False

            for file in files:
                file_path = os.path.join(root, file)

                # Processa arquivos compactados
                if file.lower().endswith('.zip') or file.lower().endswith('.rar'):
                    sub_destino = os.path.join(root, os.path.splitext(file)[0])
                    os.makedirs(sub_destino, exist_ok=True)
                    try:
                        if file.lower().endswith('.zip'):
                            with zipfile.ZipFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                        elif file.lower().endswith('.rar'):
                            with rarfile.RarFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                        os.remove(file_path)
                    except Exception as e:
                        self.registrar_erro(file_path, e)

                # Cria objetos específicos para os arquivos e os adiciona ao grupo
                elif file.upper().endswith(".PDF"):
                    pdf_obj = PDF(file_path, os.path.basename(file))
                    self.pdfs.append(pdf_obj)
                    grupo.adicionar_arquivo(pdf_obj)
                    arquivos_encontrados = True
                elif file.upper().endswith(".DXF"):
                    dxf_obj = DXF(file_path, os.path.basename(file))
                    self.dxfs.append(dxf_obj)
                    grupo.adicionar_arquivo(dxf_obj)
                    arquivos_encontrados = True
                elif file.upper().endswith(".DWG"):
                    dwg_obj = DWG(file_path, os.path.basename(file))
                    self.dwgs.append(dwg_obj)
                    grupo.adicionar_arquivo(dwg_obj)
                    arquivos_encontrados = True

            # Adiciona o grupo se encontrar arquivos relevantes
            if arquivos_encontrados:
                self.grupos.append(grupo)

    def registrar_erro(self, file_path, erro):
            """Registra erros em um arquivo de log."""
            with open("relatorio.txt", "a") as f:
                f.write(f"Erro ao processar o arquivo {file_path}:\n{str(erro)}\n")

    def create_sheet(self):
  
        pdf_files_used = set()

        # Nome da planilha
        sheet_name = (os.path.splitext(self.folder)[0]) + '.xlsx'
        exist_file = os.path.exists(sheet_name)

        # Se o arquivo já existir, carregue-o
        if exist_file:
            wb = openpyxl.load_workbook(sheet_name)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dados"
            # Cabeçalho da planilha
            
            ws.append(["ORDEM DE COMPRA", "DATA", "CÓD. PROTHEUS", "CÓD. PEÇA", "DESCRIÇÃO", "QTD", "CÓD. CHAPA", "ESPESSURA", "MATERIAL", 
                    "COMPRIMENTO", "LARGURA", "ÁREA TOTAL", "AREA SUPERFICIAL", "TEMPO DE CORTE", "DOBRAS","REBITES", 
                    "QTD REBITES", "ROSCAS","LOTE MÍNIMO (30% DA CHAPA)", "QTD CHAPA TOTAL", "ARQUIVO PDF","NOME DO PDF", "INFORME O CAMINHO ATÉ A PASTA", "EXEMPLO"])
            ws.append(["", "" ,"" ,"" ,"" , "" ,"", "","", "", "","","","","","","", "", "","","", "", "mathe\OneDrive\Área de Trabalho", "SeuUsuario\Área de trabalho\ "])
        # Preenche os dados
        for grupo in self.grupos:
            for dxf in grupo.dxfs:
                for pdf in grupo.pdfs:
                    if pdf.pdf_file in pdf_files_used:
                        continue
                    if pdf.name in dxf.dxf_name:
                        grupo.rows_to_add.append(["OC033312","20/01/2025", " ", pdf.name, pdf.protheus, " ", pdf.code, pdf.espessura, pdf.material, 
                                                dxf.comprimento, dxf.largura, dxf.area, pdf.area, dxf.cut_time, pdf.dobras, pdf.rebites,
                                                pdf.qtd_rebites, pdf.rosca, dxf.lote_min, dxf.lote_max, f'=HIPERLINK("C:\\Users\\" & T2 & "\\{pdf.pdf_local}", "Abrir PDF - {pdf.name}")', pdf.pdf_name])
                        pdf_files_used.add(pdf.pdf_file)
                        break

            # Adicionando as linhas do grupo
            for pdf in grupo.pdfs:
                if pdf.pdf_file not in pdf_files_used:
                    grupo.rows_to_add.append(["OC033312","20/01/2025", " ", pdf.name, pdf.protheus, " ", pdf.code, pdf.espessura, " ", 
                                                " ", " ", " ", pdf.area, " ", pdf.dobras, pdf.rebites,
                                                pdf.qtd_rebites, pdf.rosca, " ", " ", f'=HIPERLINK("C:\\Users\\" & T2 & "\\{pdf.pdf_local}", "Abrir PDF - {pdf.name}")', pdf.pdf_name])
                    pdf_files_used.add(pdf.pdf_file)

            grupo.rows_to_add.append([])

        # Inserindo os dados e as fórmulas dinamicamente
        for grupo in self.grupos:
            for row in grupo.rows_to_add_first:
                ws.append(row)
            for row in grupo.rows_to_add:
                ws.append(row)
      
        # Salva a planilha
        wb.save(sheet_name)

        return sheet_name
    
# Cria pastas de acordo o material de cada projeto e os move de uma forma que todos os arquivos do mesmo tipo fiquem agrupados de acordo o seu material
    def organize_folders(self):
        
        def create_and_move_file(file_path, file_name, file_src):
            try:
                # Cria a pasta se não existir
                if not os.path.exists(file_path):
                    os.makedirs(file_path)
                    
                    # Dá permissão total (rwx) para todos os grupos (777)
                    os.chmod(file_path, 0o777)
                    
                destino_path = os.path.join(file_path, file_name)
                
                # Remove o arquivo de destino se ele já existir
                if os.path.exists(destino_path):
                    os.remove(destino_path)
                
                # Move o arquivo para o destino
                shutil.move(file_src, destino_path)

                # Dá permissão total (rwx) para todos os grupos (777) no arquivo movido
                os.chmod(destino_path, 0o777)
                
            except OSError as e:
                print(f"Erro ao criar ou mover o arquivo {file_name}: {e}")
            except Exception as e:
                print(f"Ocorreu um erro inesperado ao mover {file_name}: {e}")
                
        def create_and_copy_file(file_path, file_name, file_src):
            try:
                # Cria a pasta se não existir
                if not os.path.exists(file_path):
                    os.makedirs(file_path)
                    
                    # Dá permissão total (rwx) para todos os grupos (777)
                    os.chmod(file_path, 0o777)
                    
                destino_path = os.path.join(file_path, file_name)
                
                # Remove o arquivo de destino se ele já existir
                if os.path.exists(destino_path):
                    os.remove(destino_path)
                
                # Move o arquivo para o destino
                shutil.copy(file_src, destino_path)

                # Dá permissão total (rwx) para todos os grupos (777) no arquivo movido
                os.chmod(destino_path, 0o777)
                
            except OSError as e:
                print(f"Erro ao criar ou mover o arquivo {file_name}: {e}")
            except Exception as e:
                print(f"Ocorreu um erro inesperado ao mover {file_name}: {e}")

        try:
           
            # Cria as pastas caso não existam
            for directory in [self.all_pdfs, self.all_dxfs, self.all_dwgs]:
                if os.path.exists(directory):
                    shutil.rmtree(directory)
                
                os.makedirs(directory)
                    
                # Dá permissão total (rwx) para todos os grupos (777)
                os.chmod(directory, 0o777)

            # Organiza os PDFs
            for pdf in self.pdfs:
                pdf_path = os.path.join(self.all_pdfs, pdf.material)
                create_and_copy_file(pdf_path, pdf.pdf_name, pdf.pdf_file)
                
                # Associa o material dos PDFs aos DXFs e DWGs correspondentes
                for dxf in self.dxfs:
                    if pdf.name in dxf.dxf_name:
                        dxf.material = pdf.material
                
                for dwg in self.dwgs:
                    if pdf.name in dwg.dwg_name:
                        dwg.material = pdf.material

            # Organiza os DXFs
            for dxf in self.dxfs:
                dxf_path = os.path.join(self.all_dxfs, dxf.material)
                create_and_move_file(dxf_path, dxf.dxf_name, dxf.dxf_file)  

            # Organiza os DWGs
            for dwg in self.dwgs:
                dwg_path = os.path.join(self.all_dwgs, dwg.material)
                create_and_move_file(dwg_path, dwg.dwg_name, dwg.dwg_file)

        except OSError as e:
            print(f"Erro ao criar pastas ou mover arquivos: {e}")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
        
        for pdf in self.pdfs:
            print(pdf.pdf_file)

# Realiza a compactação das pastas criadas                   
    def zip_file_process(self):
        
        folder_zip = "DESENHOS PDFs.zip"
        all_zip = "ProcessedFiles.zip"
        relatorio = "relatorio.txt"
        pdfs_zip = "PDFs.zip"
        dxfs_zip = "DXFs.zip"
        dwgs_zip = "DWGs.zip"
    
        def zip_folder (base_file, zip_file):
      
            with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root, dirs, files in os.walk(base_file):
                            for file in files:
                                caminho = os.path.join(root, file)
                                zipf.write(caminho, os.path.relpath(caminho, base_file))
                
            return zip_file
                                        
        zip_pdfs = zip_folder(self.all_pdfs, pdfs_zip)
        zip_dxfs = zip_folder(self.all_dxfs, dxfs_zip)
        zip_dwgs = zip_folder(self.all_dwgs, dwgs_zip)
        zip_extraction_folder =  zip_folder(self.extraction_folder, folder_zip)
        
        if os.path.exists(self.organization_folder_zip):
            shutil.rmtree(self.organization_folder_zip)
        
        os.makedirs(self.organization_folder_zip) 
        
        shutil.move(zip_pdfs, self.organization_folder_zip)
        shutil.move(zip_dxfs, self.organization_folder_zip)
        shutil.move(zip_dwgs, self.organization_folder_zip)
        shutil.move(self.sheets, self.organization_folder_zip)
        shutil.move(zip_extraction_folder, self.organization_folder_zip)
        
        if os.path.exists(relatorio):
            shutil.move(relatorio, self.organization_folder_zip)
            
        zip_file_processed = zip_folder(self.organization_folder_zip, all_zip)
        
        destino_path = os.path.join("uploads", zip_file_processed)
        
        if os.path.exists(destino_path):
            os.remove(destino_path)
        
        shutil.move(zip_file_processed, "uploads")

        return zip_file_processed

# Apaga pastas criadas durante o processo    
    def clean_all(self):
        
        if os.path.exists(self.extraction_folder):
            shutil.rmtree(self.extraction_folder)
        if os.path.exists(self.organization_folder):
            shutil.rmtree(self.organization_folder)
        if os.path.exists(self.organization_folder_zip):
            shutil.rmtree(self.organization_folder_zip)
        if os.path.exists(self.folder):
            os.remove(self.folder)    

def extract_subfiles(root_folder):
        
        processed_files = set()
        stack = [root_folder]

        while stack:
            current_folder = stack.pop()

            for root, dirs, files in os.walk(current_folder):
                for file in files:
                    file_path = os.path.join(root, file)

                    if file_path in processed_files:
                        continue

                    processed_files.add(file_path)

                    # Tratamento de erro para arquivos ZIP
                    if file.lower().endswith('.zip'):
                        sub_destino = os.path.join(root, os.path.splitext(file)[0])
                        os.makedirs(sub_destino, exist_ok=True)
                        try:
                            with zipfile.ZipFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                            os.remove(file_path)
                            stack.append(sub_destino)
                        except zipfile.BadZipFile as e:
                            with open("relatorio.txt", "a") as f:
                                f.write(f"O arquivo {file_path} não é um ZIP válido ou está corrompido.\n")
                                f.write(f"Erro ao processar o arquivo {file_path}:\n")
                                f.write(f"{str(e)}\n")

                    # Tratamento de erro para arquivos RAR
                    elif file.lower().endswith('.rar'):
                        sub_destino = os.path.join(root, os.path.splitext(file)[0])
                        os.makedirs(sub_destino, exist_ok=True)
                        try:
                            with rarfile.RarFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                            os.remove(file_path)
                            stack.append(sub_destino)
                        except rarfile.Error as e:
                            with open("relatorio.txt", "a") as f:
                                f.write(f"O arquivo {file_path} não é um RAR válido ou está corrompido.\n")
                                f.write(f"Erro ao processar o arquivo {file_path}:\n")
                                f.write(f"{str(e)}\n")

        return root_folder  # Retorna a pasta raiz que contém tudo
